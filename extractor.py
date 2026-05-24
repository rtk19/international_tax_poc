from __future__ import annotations

import argparse
import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class TableConfig:
    table_name: str
    headers: list[str] | None = None
    header_starts_with: str | None = None


@dataclass
class SheetConfig:
    name: str
    key_value_section: dict[str, Any] | None
    tables: list[TableConfig]


@dataclass
class FileTypeConfig:
    name: str
    filename_contains: str
    sheets: list[SheetConfig]


class ExtractionError(Exception):
    pass


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_config(config_path: Path) -> list[FileTypeConfig]:
    raw = json.loads(config_path.read_text(encoding="utf-8"))
    file_types: list[FileTypeConfig] = []
    for file_type in raw.get("file_types", []):
        sheets: list[SheetConfig] = []
        for sheet in file_type.get("sheets", []):
            tables = [
                TableConfig(
                    table_name=t["table_name"],
                    headers=t.get("headers"),
                    header_starts_with=t.get("header_starts_with"),
                )
                for t in sheet.get("tables", [])
            ]
            sheets.append(
                SheetConfig(
                    name=sheet["name"],
                    key_value_section=sheet.get("key_value_section"),
                    tables=tables,
                )
            )

        file_types.append(
            FileTypeConfig(
                name=file_type["name"],
                filename_contains=file_type["filename_contains"],
                sheets=sheets,
            )
        )

    if not file_types:
        raise ExtractionError("No file_types found in configuration.")

    return file_types


def detect_file_type(file_path: Path, file_types: list[FileTypeConfig]) -> FileTypeConfig | None:
    filename = file_path.name
    for file_type in file_types:
        if file_type.filename_contains in filename:
            return file_type
    return None


def find_header_row_by_exact_headers(sheet: Worksheet, headers: list[str]) -> tuple[int, list[int]]:
    headers_norm = [h.strip() for h in headers]
    max_row, max_col = sheet.max_row, sheet.max_column

    for row in range(1, max_row + 1):
        row_values = [normalize_cell(sheet.cell(row=row, column=col).value) for col in range(1, max_col + 1)]
        for start_idx in range(0, max_col - len(headers_norm) + 1):
            segment = row_values[start_idx : start_idx + len(headers_norm)]
            if segment == headers_norm:
                col_indexes = list(range(start_idx + 1, start_idx + len(headers_norm) + 1))
                return row, col_indexes

    raise ExtractionError(f"Could not find table headers: {headers}")


def find_header_row_by_starts_with(sheet: Worksheet, header_prefix: str) -> tuple[int, int]:
    max_row, max_col = sheet.max_row, sheet.max_column
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            value = normalize_cell(sheet.cell(row=row, column=col).value)
            if value.startswith(header_prefix):
                return row, col
    raise ExtractionError(f"Could not find table header starting with: {header_prefix}")


def extract_key_values(sheet: Worksheet, key_col: str, value_col: str, start_row: int) -> dict[str, Any]:
    extracted: dict[str, Any] = {}
    row = start_row

    while True:
        key = normalize_cell(sheet[f"{key_col}{row}"].value)
        value = sheet[f"{value_col}{row}"].value

        if key == "":
            break

        extracted[key] = value
        row += 1

    return extracted


def extract_table(sheet: Worksheet, header_row: int, col_indexes: list[int], headers: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    row = header_row + 1

    while True:
        values = [sheet.cell(row=row, column=col).value for col in col_indexes]
        if all(normalize_cell(v) == "" for v in values):
            break

        rows.append({header: value for header, value in zip(headers, values, strict=False)})
        row += 1

    return rows


def extract_dynamic_year_table(sheet: Worksheet, header_row: int, start_col: int) -> list[dict[str, Any]]:
    headers: list[str] = []
    col = start_col

    while True:
        value = normalize_cell(sheet.cell(row=header_row, column=col).value)
        if value == "":
            break
        headers.append(value)
        col += 1

    col_indexes = list(range(start_col, start_col + len(headers)))
    return extract_table(sheet, header_row, col_indexes, headers)


def extract_from_sheet(sheet: Worksheet, sheet_config: SheetConfig) -> dict[str, Any]:
    result: dict[str, Any] = {}

    if sheet_config.key_value_section:
        result["key_values"] = extract_key_values(
            sheet,
            key_col=sheet_config.key_value_section["key_column"],
            value_col=sheet_config.key_value_section["value_column"],
            start_row=sheet_config.key_value_section.get("start_row", 1),
        )

    for table_config in sheet_config.tables:
        if table_config.headers:
            header_row, cols = find_header_row_by_exact_headers(sheet, table_config.headers)
            result[table_config.table_name] = extract_table(sheet, header_row, cols, table_config.headers)
        elif table_config.header_starts_with:
            header_row, start_col = find_header_row_by_starts_with(sheet, table_config.header_starts_with)
            result[table_config.table_name] = extract_dynamic_year_table(sheet, header_row, start_col)

    return result


def extract_file(file_path: Path, file_type: FileTypeConfig) -> dict[str, Any]:
    wb = load_workbook(file_path, data_only=True)
    output: dict[str, Any] = {"file_name": file_path.name, "file_type": file_type.name, "sheets": {}}

    for sheet_cfg in file_type.sheets:
        if sheet_cfg.name not in wb.sheetnames:
            raise ExtractionError(f"Sheet '{sheet_cfg.name}' not found in {file_path.name}")

        sheet = wb[sheet_cfg.name]
        output["sheets"][sheet_cfg.name] = extract_from_sheet(sheet, sheet_cfg)

    return output


def write_csv_tables(extracted: dict[str, Any], csv_dir: Path) -> None:
    csv_dir.mkdir(parents=True, exist_ok=True)

    file_name = Path(extracted["file_name"]).stem
    for sheet_name, sheet_payload in extracted["sheets"].items():
        for section_name, section_data in sheet_payload.items():
            if isinstance(section_data, list) and section_data:
                fieldnames = list(section_data[0].keys())
                target = csv_dir / f"{file_name}__{sheet_name}__{section_name}.csv"
                with target.open("w", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(section_data)


def iter_xlsx_files(input_dir: Path) -> Iterable[Path]:
    return sorted(p for p in input_dir.glob("*.xlsx") if p.is_file())


def main() -> None:
    parser = argparse.ArgumentParser(description="Dynamic XLSX extractor to JSON/CSV.")
    parser.add_argument("--input-dir", default="./input", help="Folder containing XLSX files")
    parser.add_argument("--config", default="./extractor_config.json", help="Extraction config JSON path")
    parser.add_argument("--output-json", default="./output/extracted.json", help="Output JSON file")
    parser.add_argument("--output-csv-dir", default="./output/csv", help="Output CSV directory")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    output_json = Path(args.output_json)
    output_csv_dir = Path(args.output_csv_dir)

    file_types = load_config(Path(args.config))

    all_results: list[dict[str, Any]] = []
    for xlsx in iter_xlsx_files(input_dir):
        file_type = detect_file_type(xlsx, file_types)
        if not file_type:
            continue
        extracted = extract_file(xlsx, file_type)
        write_csv_tables(extracted, output_csv_dir)
        all_results.append(extracted)

    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(all_results, indent=2, ensure_ascii=False, default=str), encoding="utf-8")

    print(f"Processed files: {len(all_results)}")
    print(f"JSON output: {output_json}")
    print(f"CSV output dir: {output_csv_dir}")


if __name__ == "__main__":
    main()
